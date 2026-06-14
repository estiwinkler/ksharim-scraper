"""
קשרים בריבוע - Scraper חכם עם Groq AI
- מנסה לפתור עם AI (עד 4 טעויות)
- אחרי 4 טעויות: טועה בכוונה פעם 5 → האתר מציג פתרון → אוסף נתונים
- שומר JSON ו-Excel מצטברים (לא מוחק ימים קודמים)
"""

import asyncio
import json
import re
import os
import urllib.request
from datetime import date, datetime
from pathlib import Path

from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# הגדרות
# ============================================================
GROQ_API_KEY   = os.environ.get("GROQ_API_KEY", "הכנסי את ה-KEY שלך כאן")
GROQ_MODEL     = "llama-3.3-70b-versatile"
URL            = "https://ksharim-baribua.com/"
MAX_WRONG_GUESSES = 4   # אחרי 4 טעויות → בזבז את הניחוש ה-5 כדי לחשוף פתרון

DATA_DIR   = Path("data")
DATA_DIR.mkdir(exist_ok=True)
JSON_FILE  = DATA_DIR / "puzzles_database.json"
EXCEL_FILE = DATA_DIR / "puzzles_database.xlsx"
LOG_FILE   = DATA_DIR / "scrape_log.txt"


# ============================================================
# לוג
# ============================================================
def log(message: str):
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {message}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ============================================================
# Groq AI
# ============================================================
def ask_groq(prompt: str) -> str:
    body = json.dumps({
        "model": GROQ_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 500,
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.groq.com/openai/v1/chat/completions",
        data=body,
        headers={
            "Content-Type":  "application/json",
            "Authorization": f"Bearer {GROQ_API_KEY}",
        },
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        result = json.loads(resp.read().decode("utf-8"))
    return result["choices"][0]["message"]["content"].strip()


def guess_group(remaining: list[str], failed: list[list[str]]) -> dict:
    failed_str = ""
    if failed:
        failed_str = "\n\nקבוצות שנכשלו - אל תנחש שוב:\n"
        for f in failed:
            failed_str += f"- {', '.join(f)}\n"

    prompt = f"""אתה משחק "קשרים בריבוע" - משחק עברי שבו צריך למצוא קבוצות של 4 מילים עם קשר משותף.

המילים שנשארו: {', '.join(remaining)}
{failed_str}
בחר את הקבוצה שאתה הכי בטוח בה.

ענה ב-JSON בלבד, ללא טקסט נוסף:
{{"category": "שם הקטגוריה", "words": ["מילה1", "מילה2", "מילה3", "מילה4"]}}"""

    text = ask_groq(prompt)
    text = re.sub(r"```json|```", "", text).strip()
    return json.loads(text)


def guess_almost_fix(remaining: list[str], almost_words: list[str], failed: list[list[str]]) -> dict:
    failed_str = ""
    if failed:
        failed_str = "\n\nשילובים שנכשלו לגמרי:\n"
        for f in failed:
            failed_str += f"- {', '.join(f)}\n"

    prompt = f"""אתה משחק "קשרים בריבוע" - משחק עברי.

ניסיתי: {', '.join(almost_words)}
האתר אמר "כמעט..." - 3 מתוך 4 נכונות, אחת שגויה.

כל המילים שנשארו: {', '.join(remaining)}
{failed_str}
החלף מילה אחת שגויה במילה אחרת מהרשימה.

ענה ב-JSON בלבד:
{{"category": "שם הקטגוריה", "words": ["מילה1", "מילה2", "מילה3", "מילה4"]}}"""

    text = ask_groq(prompt)
    text = re.sub(r"```json|```", "", text).strip()
    return json.loads(text)


# ============================================================
# פעולות Playwright
# ============================================================
async def handle_popups(page):
    """סוגר חלונות פופאפ ומכניס שם אם נדרש."""
    await page.wait_for_timeout(2000)

    # הכנס שם אם יש שדה שם
    try:
        name_input = page.locator('input[type="text"], input[placeholder*="שם"], input[name*="name"]').first
        if await name_input.count() > 0:
            log("  מכניס שם למשחק...")
            await name_input.fill("Bot")
            await page.wait_for_timeout(500)

            # לחץ על כפתור "אפשר להתחיל לשחק" או דומה
            start_btn = page.locator("button:has-text('אפשר להתחיל לשחק'), button:has-text('התחל'), button:has-text('המשך')")
            if await start_btn.count() > 0:
                await start_btn.first.click()
                log("  לחץ על כפתור התחלה")
                await page.wait_for_timeout(1500)
    except Exception as e:
        log(f"  שם/התחלה: {e}")

    # סגור חלון "איך משחקים" אם פתוח
    try:
        close_btn = page.get_by_role("button", name=re.compile("סגירה|סגור|close|×|✕|המשך|OK", re.IGNORECASE))
        if await close_btn.count() > 0:
            await close_btn.first.click()
            log("  סגר חלון הוראות")
            await page.wait_for_timeout(1000)
    except Exception as e:
        log(f"  סגירת חלון: {e}")

    # לחץ ESC למקרה שיש עוד חלון
    await page.keyboard.press("Escape")
    await page.wait_for_timeout(500)

    log("  פופאפים טופלו")


async def get_words(page) -> list[str]:
    await page.wait_for_timeout(3000)

    buttons = await page.evaluate("""
        () => {
            const btns = [];
            document.querySelectorAll('button, [role="button"], [class*="word"], [class*="Word"], [class*="tile"], [class*="card"]').forEach(el => {
                const txt = el.innerText?.trim();
                if (txt && /[\u05d0-\u05ea]{2,}/.test(txt) && txt.length < 20)
                    btns.push(txt);
            });
            return [...new Set(btns)];
        }
    """)

    if len(buttons) < 12:
        raw     = await page.inner_text("body")
        hebrew  = re.findall(r'[\u05d0-\u05ea]{2,}', raw)
        buttons = list(dict.fromkeys(hebrew))

    words = [w for w in buttons if 1 < len(w) < 15][:16]
    log(f"נמצאו {len(words)} מילים: {', '.join(words)}")
    return words


async def click_words(page, words: list[str]) -> int:
    clicked = 0
    for word in words:
        try:
            btn = page.get_by_role("button", name=re.compile(f"^{re.escape(word)}$"))
            if await btn.count() > 0:
                await btn.first.click()
                clicked += 1
                await page.wait_for_timeout(400)
                continue
            el = page.locator(f"text={word}").first
            if await el.count() > 0:
                await el.click()
                clicked += 1
                await page.wait_for_timeout(400)
        except Exception:
            pass
    return clicked


async def deselect_all(page):
    try:
        clear = page.get_by_role("button", name=re.compile("נקה|ביטול|clear|deselect", re.IGNORECASE))
        if await clear.count() > 0:
            await clear.first.click()
            await page.wait_for_timeout(500)
            return
    except Exception:
        pass
    try:
        selected = await page.evaluate("""
            () => {
                const sel = [];
                document.querySelectorAll('[class*="selected"], [class*="active"], [aria-pressed="true"]').forEach(el => {
                    const txt = el.innerText?.trim();
                    if (txt) sel.push(txt);
                });
                return sel;
            }
        """)
        for word in selected:
            try:
                el = page.locator(f"text={word}").first
                if await el.count() > 0:
                    await el.click()
                    await page.wait_for_timeout(300)
            except Exception:
                pass
    except Exception:
        pass


async def submit_guess(page):
    """לוחץ על כפתור בדוק."""
    try:
        check = page.get_by_role("button", name=re.compile("בדוק|אשר|שלח|submit|confirm", re.IGNORECASE))
        if await check.count() > 0:
            await check.first.click()
            await page.wait_for_timeout(2000)
    except Exception:
        pass


async def close_alert(page):
    """סוגר alert/popup אם קיים."""
    try:
        close = page.get_by_role("button", name=re.compile("סגור|אישור|המשך|ok|close|×|✕", re.IGNORECASE))
        if await close.count() > 0:
            await close.first.click()
            await page.wait_for_timeout(800)
            return
        # נסה ESC
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(500)
    except Exception:
        pass


async def check_result(page) -> str:
    """
    מזהה תוצאה אחרי לחיצת בדוק.
    מחזיר: 'correct' / 'almost' / 'wrong' / 'unknown'
    """
    alert_text = await page.evaluate("""
        () => {
            const els = document.querySelectorAll('[role="alert"], [class*="alert"], [class*="modal"], [class*="popup"], [class*="toast"], [class*="message"]');
            for (const el of els) {
                const txt = el.innerText?.trim();
                if (txt && txt.length > 0) return txt;
            }
            return "";
        }
    """)

    body_result = await page.evaluate("""
        () => {
            const body = document.body.innerText;
            if (body.includes('כמעט'))                                 return 'almost';
            if (body.includes('כל הכבוד') || body.includes('ניצחת') ||
                body.includes('מעולה')    || body.includes('correct')) return 'correct';
            if (body.includes('טעות')    || body.includes('נסה שוב') ||
                body.includes('לא נכון') || body.includes('חבל')     ||
                body.includes('wrong'))                                return 'wrong';
            return 'unknown';
        }
    """)

    if alert_text:
        log(f"  אלרט: '{alert_text}'")
        if "כמעט"    in alert_text: result = "almost"
        elif any(w in alert_text for w in ["כל הכבוד","ניצחת","מעולה","correct"]): result = "correct"
        elif any(w in alert_text for w in ["טעות","נסה שוב","לא נכון","חבל","wrong"]): result = "wrong"
        else: result = body_result
    else:
        result = body_result

    log(f"  תוצאה: {result}")
    return result


# ============================================================
# איסוף פתרון מהמסך הסופי
# ============================================================
async def collect_solution_from_screen(page) -> list[dict]:
    """
    אוסף את הפתרון המוצג על המסך.
    המבנה: כותרת קטגוריה + מילים מופרדות ב-•
    לדוגמה:
      נשיאי ארצות הברית
      בוש • לינקולן • פורד • קרטר
    """
    await page.wait_for_timeout(2000)

    groups = await page.evaluate("""
        () => {
            const results = [];
            // חפש בלוקים של קטגוריות - לפי הצבעים/קלאסים
            const selectors = [
                '[class*="group"]', '[class*="Group"]',
                '[class*="category"]', '[class*="Category"]',
                '[class*="solved"]', '[class*="result"]',
                '[class*="card"]'
            ];

            for (const sel of selectors) {
                document.querySelectorAll(sel).forEach(el => {
                    const txt = el.innerText?.trim();
                    if (!txt || txt.length < 5) return;

                    // בדוק אם יש • (מפריד בין מילים)
                    if (txt.includes('•')) {
                        const lines = txt.split('\\n').map(l => l.trim()).filter(Boolean);
                        if (lines.length >= 2) {
                            const category = lines[0];
                            const wordsLine = lines.find(l => l.includes('•'));
                            if (wordsLine) {
                                const words = wordsLine.split('•').map(w => w.trim()).filter(Boolean);
                                if (words.length >= 3) {
                                    results.push({ category, words });
                                }
                            }
                        }
                    }
                });
                if (results.length >= 4) break;
            }
            return results;
        }
    """)

    # אם לא מצא, נסה גישה גנרית לפי טקסט הגוף
    if len(groups) < 2:
        log("  מנסה גישה חלופית לאיסוף פתרון...")
        raw_text = await page.inner_text("body")
        groups = parse_solution_from_text(raw_text)

    log(f"  נאספו {len(groups)} קבוצות מהמסך")
    for g in groups:
        log(f"    📌 {g.get('category')}: {', '.join(g.get('words', []))}")

    return groups


def parse_solution_from_text(text: str) -> list[dict]:
    """מנתח טקסט גולמי ומחפש קטגוריות ומילים."""
    groups = []
    lines  = [l.strip() for l in text.split('\n') if l.strip()]

    for i, line in enumerate(lines):
        if '•' in line:
            # השורה הזו היא מילים - הקטגוריה בשורה לפניה
            words = [w.strip() for w in line.split('•') if w.strip()]
            if len(words) >= 3:
                category = lines[i-1] if i > 0 else "קטגוריה לא ידועה"
                # וודא שהקטגוריה לא מכילה • (שלא תהיה שורת מילים נוספת)
                if '•' not in category and len(category) < 50:
                    groups.append({"category": category, "words": words, "source": "screen"})

    return groups[:4]


# ============================================================
# לולאת הפתרון הראשית
# ============================================================
async def solve_puzzle(page, words_16: list[str]) -> dict:
    """
    1. מנסה לפתור עם Groq (עד MAX_WRONG_GUESSES טעויות)
    2. אם הגיע למגבלה → מבזבז את הניחוש האחרון → אוסף פתרון מהמסך
    3. אם פתר הכל נכון → שומר את הפתרון
    """
    solved_groups = []    # קבוצות שנפתרו נכון
    remaining     = list(words_16)
    failed        = []
    almost_words  = None
    wrong_count   = 0     # סופר טעויות
    attempt       = 0
    max_attempts  = 15

    while len(solved_groups) < 4 and attempt < max_attempts and len(remaining) >= 4:
        attempt += 1
        log(f"\n--- ניסיון {attempt} | נפתרו: {len(solved_groups)}/4 | טעויות: {wrong_count}/{MAX_WRONG_GUESSES} ---")

        # הגענו למגבלת הטעויות → בזבז ניחושים עד שהפתרון מתגלה
        if wrong_count >= MAX_WRONG_GUESSES:
            log("⚠️ הגענו ל-4 טעויות → מבזבז ניחושים לחשיפת פתרון...")
            import random

            # בצע עד 5 ניחושים רנדומליים כדי לרוקן את כל הניסיונות
            shuffled = list(remaining)
            random.shuffle(shuffled)
            for attempt_idx in range(5):
                guess_words = shuffled[attempt_idx*4 % len(shuffled):(attempt_idx*4 % len(shuffled))+4]
                if len(guess_words) < 4:
                    guess_words = shuffled[:4]
                log(f"  ניחוש מבזבז {attempt_idx+1}: {guess_words}")
                await click_words(page, guess_words)
                await submit_guess(page)
                await page.wait_for_timeout(2500)
                await close_alert(page)
                await page.wait_for_timeout(1000)

                # בדוק אם הפתרון כבר מוצג (יש • על המסך)
                has_solution = await page.evaluate("""
                    () => document.body.innerText.includes('•')
                """)
                if has_solution:
                    log("  ✅ פתרון מוצג על המסך!")
                    break
            await page.wait_for_timeout(4000)
            await close_alert(page)
            await page.wait_for_timeout(2000)

            # לחץ על "הצגת תוצאות" אם קיים
            try:
                show_btn = page.get_by_role("button", name=re.compile("הצגת תוצאות|הצג תוצאות|פתרון|תוצאות|ראה פתרון", re.IGNORECASE))
                if await show_btn.count() > 0:
                    log("  לוחץ על הצגת תוצאות...")
                    await show_btn.first.click()
                    await page.wait_for_timeout(3000)
                else:
                    log("  לא נמצא כפתור הצגת תוצאות, ממתין לטעינה...")
                    await page.wait_for_timeout(4000)
            except Exception as btn_err:
                log(f"  שגיאה בחיפוש כפתור: {btn_err}")
                await page.wait_for_timeout(3000)

            # צלם מסך לדיאגנוסטיקה
            screenshot_path = DATA_DIR / f"debug_{date.today().isoformat()}.png"
            await page.screenshot(path=str(screenshot_path), full_page=True)
            log(f"  צילום מסך נשמר: {screenshot_path}")

            # שמור גם את ה-HTML המלא
            html_path = DATA_DIR / f"debug_{date.today().isoformat()}.html"
            html_content = await page.content()
            with open(html_path, 'w', encoding='utf-8') as hf:
                hf.write(html_content)
            log(f"  HTML נשמר: {html_path}")

            # שמור טקסט גולמי
            raw_text = await page.inner_text("body")
            txt_path = DATA_DIR / f"debug_{date.today().isoformat()}.txt"
            with open(txt_path, 'w', encoding='utf-8') as tf:
                tf.write(raw_text)
            log(f"  טקסט גולמי:\n{raw_text[:1000]}")

            # אסוף פתרון מהמסך
            screen_solution = await collect_solution_from_screen(page)

            # שלב: קבוצות שכבר פתרנו + קבוצות שנאספו מהמסך
            all_groups = solved_groups[:]
            solved_cats = {g["category"] for g in solved_groups}
            for g in screen_solution:
                if g.get("category") not in solved_cats:
                    all_groups.append({**g, "correct": True, "source": "screen"})

            return {
                "solved_groups":  all_groups,
                "total_solved":   len(all_groups),
                "source":         "screen_after_fail",
                "wrong_guesses":  wrong_count + 1,
            }

        # נחש עם Groq
        try:
            if almost_words:
                group = guess_almost_fix(remaining, almost_words, failed)
                log(f"Groq מתקן כמעט: {group.get('category')} → {group.get('words')}")
            else:
                group = guess_group(remaining, failed)
                log(f"Groq מנחש: {group.get('category')} → {group.get('words')}")
        except Exception as e:
            log(f"שגיאת Groq: {e} → עובר למצב ניחושים אוטומטי")
            wrong_count = MAX_WRONG_GUESSES
            continue

        cat   = group.get("category", "")
        words = [w for w in group.get("words", []) if w in remaining]

        if len(words) != 4:
            log(f"  ⚠️ מילים לא תקינות")
            failed.append(group.get("words", []))
            almost_words = None
            continue

        clicked = await click_words(page, words)
        log(f"  לחצתי על {clicked}/4 מילים")
        await submit_guess(page)
        result = await check_result(page)
        await close_alert(page)

        if result == "correct":
            log(f"  ✅ נכון! {cat}")
            solved_groups.append({"category": cat, "words": words, "correct": True, "source": "ai"})
            remaining    = [w for w in remaining if w not in words]
            failed       = []
            almost_words = None

        elif result == "almost":
            log(f"  🟡 כמעט! מחליף מילה אחת")
            almost_words = words
            await deselect_all(page)

        else:
            log(f"  ❌ טעות ({wrong_count + 1}/{MAX_WRONG_GUESSES})")
            failed.append(words)
            almost_words = None
            wrong_count += 1
            await deselect_all(page)

    # פתר הכל נכון!
    if len(solved_groups) == 3 and len(remaining) == 4:
        solved_groups.append({"category": "קבוצה אחרונה", "words": remaining, "correct": True, "source": "ai"})

    return {
        "solved_groups": solved_groups,
        "total_solved":  len(solved_groups),
        "source":        "ai",
        "wrong_guesses": wrong_count,
    }


# ============================================================
# סריקה ראשית
# ============================================================
async def fetch_and_solve() -> dict:
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(locale="he-IL")
        page    = await context.new_page()

        await page.goto(URL, wait_until="networkidle", timeout=30000)

        # טפל בפופאפים לפני הכל
        await handle_popups(page)

        words = await get_words(page)

        if len(words) < 8:
            await browser.close()
            return {"error": "לא נמצאו מספיק מילים", "words": words, "solved_groups": []}

        log("מתחיל לפתור...")
        result = await solve_puzzle(page, words)
        await browser.close()

        result["words_found"] = words
        return result


# ============================================================
# שמירה ל-JSON (מצטבר - לא מוחק ימים קודמים)
# ============================================================
def save_to_json(puzzle: dict):
    # טען מאגר קיים או צור חדש
    if JSON_FILE.exists():
        with open(JSON_FILE, "r", encoding="utf-8") as f:
            db = json.load(f)
    else:
        db = {"puzzles": [], "total": 0, "created": date.today().isoformat()}

    today = date.today().isoformat()

    # עדכן אם היום כבר קיים, אחרת הוסף
    existing = [i for i, p in enumerate(db["puzzles"]) if p.get("date") == today]
    if existing:
        db["puzzles"][existing[0]] = puzzle
        log(f"  עודכן רשומה קיימת ליום {today}")
    else:
        db["puzzles"].append(puzzle)
        log(f"  נוספה רשומה חדשה ליום {today}")

    db["total"]        = len(db["puzzles"])
    db["last_updated"] = datetime.now().isoformat()

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
    log(f"✅ JSON נשמר - סה\"כ {db['total']} ימים במאגר")


# ============================================================
# שמירה ל-Excel (מצטבר - לא מוחק ימים קודמים)
# ============================================================
def save_to_excel(puzzle: dict):
    HEADER_BG = "2E4057"
    COLORS    = {0: "ABEBC6", 1: "FAD7A0", 2: "F1948A", 3: "AED6F1"}

    # טען קובץ קיים או צור חדש עם כותרות
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "קשרים בריבוע"
        headers  = ["תאריך", "קטגוריה", "מילה 1", "מילה 2", "מילה 3", "מילה 4", "נכון?", "מקור"]
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=12, name="Arial")
            cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 32
        for c in ["C","D","E","F"]: ws.column_dimensions[c].width = 14
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 10
        ws.row_dimensions[1].height = 28

    today  = puzzle.get("date", date.today().isoformat())
    groups = puzzle.get("solved_groups", [])

    # מחק שורות של היום אם קיימות (עדכון)
    rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        if row[0].value == today:
            rows_to_delete.append(row[0].row)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)

    # הוסף שורות חדשות
    for i, group in enumerate(groups):
        row   = ws.max_row + 1
        words = group.get("words", [])
        src   = "🤖 AI" if group.get("source") == "ai" else "📺 מסך"
        vals  = [
            today,
            group.get("category", ""),
            *words[:4],
            "✅" if group.get("correct") else "❓",
            src,
        ]
        for col, val in enumerate(vals, 1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=COLORS.get(i % 4, "FFFFFF"))
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font      = Font(name="Arial", size=11)
            cell.border    = Border(
                bottom=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin",  color="CCCCCC")
            )
        ws.row_dimensions[row].height = 24

    # שורה ריקה בין ימים
    ws.append([""])
    wb.save(EXCEL_FILE)
    log(f"✅ Excel נשמר - נוספו {len(groups)} קבוצות")


# ============================================================
# Main
# ============================================================
async def main():
    log("=" * 50)
    log(f"סריקה: {date.today().isoformat()}")

    result               = await fetch_and_solve()
    result["date"]       = date.today().isoformat()
    result["fetched_at"] = datetime.now().isoformat()

    save_to_json(result)
    save_to_excel(result)

    log("=" * 50)
    log("📋 תוצאות סופיות:")
    for g in result.get("solved_groups", []):
        src    = "🤖" if g.get("source") == "ai" else "📺"
        status = "✅" if g.get("correct") else "❓"
        log(f"  {status}{src} {g.get('category')}: {', '.join(g.get('words', []))}")
    log(f"סה\"כ: {result.get('total_solved', 0)}/4 | טעויות: {result.get('wrong_guesses', 0)}")
    log(f"מקור: {result.get('source', 'unknown')}")
    log("✅ הושלם!")


if __name__ == "__main__":
    asyncio.run(main())
